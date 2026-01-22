#!/usr/bin/env python3
"""
Creates the complete Unlock Egypt content Excel file with all 5 sheets.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============ SHEET 1: Sites ============
ws_sites = wb.active
ws_sites.title = "Sites"

sites_headers = ["id", "name", "arabicName", "era", "tourismType", "placeType", "city",
                 "shortDescription", "latitude", "longitude", "imageUrl", "isUnlocked",
                 "estimatedDuration", "bestTimeToVisit"]
ws_sites.append(sites_headers)

sites_data = [
    ["giza", "Pyramids of Giza", "أهرامات الجيزة", "Old Kingdom", "Pharaonic", "Pyramid", "Giza",
     "The last surviving wonder of the ancient world.", 29.9792, 31.1342, "", "TRUE", "3-4 hours", "Early morning (8-10 AM) or late afternoon"],
    ["luxor", "Luxor Temple", "معبد الأقصر", "New Kingdom", "Pharaonic", "Temple", "Luxor",
     "A stunning temple in the heart of modern Luxor, beautiful at night.", 25.6996, 32.639, "", "TRUE", "2-3 hours", "Evening visit recommended - beautifully lit at night"],
    ["valley_kings", "Valley of the Kings", "وادي الملوك", "New Kingdom", "Pharaonic", "Tomb", "Luxor",
     "The hidden burial ground of Egypt's greatest pharaohs.", 25.7402, 32.6014, "", "TRUE", "3-4 hours", "Early morning (6-8 AM) before the heat"],
    ["karnak", "Karnak Temple Complex", "معابد الكرنك", "New Kingdom", "Pharaonic", "Temple", "Luxor",
     "The largest ancient religious complex ever built.", 25.7188, 32.6573, "", "TRUE", "3-4 hours", "Early morning or evening Sound & Light show"],
    ["abu_simbel", "Abu Simbel", "أبو سمبل", "New Kingdom", "Pharaonic", "Temple", "Aswan",
     "Ramesses II's monumental temple, famously relocated.", 22.3372, 31.6258, "", "TRUE", "2-3 hours", "Feb 22 or Oct 22 for the sun festival"],
]

for row in sites_data:
    ws_sites.append(row)

style_header(ws_sites, len(sites_headers))
auto_width(ws_sites)

# Add dropdowns for Sites sheet
era_values = '"Pre-Dynastic,Old Kingdom,Middle Kingdom,New Kingdom,Late Period,Ptolemaic,Roman,Islamic,Modern"'
tourism_values = '"Pharaonic,Greco-Roman,Coptic,Islamic,Modern"'
place_values = '"Pyramid,Temple,Tomb,Museum,Mosque,Church,Fortress,Market,Monument,Ruins"'
city_values = '"Cairo,Giza,Luxor,Aswan,Alexandria,Sinai,Fayoum,Dahab,Hurghada,Sharm El Sheikh"'

dv_era = DataValidation(type="list", formula1=era_values, allow_blank=True)
dv_tourism = DataValidation(type="list", formula1=tourism_values, allow_blank=True)
dv_place = DataValidation(type="list", formula1=place_values, allow_blank=True)
dv_city = DataValidation(type="list", formula1=city_values, allow_blank=True)

ws_sites.add_data_validation(dv_era)
ws_sites.add_data_validation(dv_tourism)
ws_sites.add_data_validation(dv_place)
ws_sites.add_data_validation(dv_city)

dv_era.add('D2:D100')
dv_tourism.add('E2:E100')
dv_place.add('F2:F100')
dv_city.add('G2:G100')

# ============ SHEET 2: SubLocations ============
ws_subloc = wb.create_sheet("SubLocations")

subloc_headers = ["id", "siteId", "name", "arabicName", "shortDescription", "imageUrl", "order"]
ws_subloc.append(subloc_headers)

subloc_data = [
    ["great_pyramid", "giza", "Great Pyramid of Khufu", "هرم خوفو الأكبر", "The largest and oldest of the three pyramids", "", 1],
    ["sphinx", "giza", "The Great Sphinx", "أبو الهول", "The mysterious guardian with a lion's body and human head", "", 2],
    ["khafre_pyramid", "giza", "Pyramid of Khafre", "هرم خفرع", "The second-largest pyramid, built for Khufu's son", "", 3],
    ["solar_boat", "giza", "Solar Boat Museum", "متحف مركب الشمس", "Houses a reconstructed ancient boat buried near the Great Pyramid", "", 4],
    ["luxor_entrance", "luxor", "The Grand Entrance", "المدخل الكبير", "Massive pylons and the remaining obelisk", "", 1],
    ["tut_tomb", "valley_kings", "Tomb of Tutankhamun", "مقبرة توت عنخ آمون", "The famous tomb of the boy king", "", 1],
    ["hypostyle_hall", "karnak", "Great Hypostyle Hall", "قاعة الأعمدة الكبرى", "134 massive columns in a forest of stone", "", 1],
    ["great_temple", "abu_simbel", "Great Temple of Ramesses II", "معبد رمسيس الثاني الكبير", "Four colossal statues guard the entrance", "", 1],
]

for row in subloc_data:
    ws_subloc.append(row)

style_header(ws_subloc, len(subloc_headers))
auto_width(ws_subloc)

# ============ SHEET 3: Cards ============
ws_cards = wb.create_sheet("Cards")

cards_headers = ["id", "subLocationId", "order", "type", "imageUrl", "content", "funFact",
                 "quizQuestion", "quizOption1", "quizOption2", "quizOption3", "quizOption4",
                 "quizCorrectAnswer", "quizExplanation"]
ws_cards.append(cards_headers)

cards_data = [
    # Great Pyramid cards
    ["khufu_1", "great_pyramid", 1, "intro", "", "Standing 481 feet tall, the Great Pyramid was the tallest structure on Earth for over 3,800 years.", "", "", "", "", "", "", "", ""],
    ["khufu_2", "great_pyramid", 2, "story", "", "Built around 2560 BCE for Pharaoh Khufu. It took roughly 20 years and 100,000 workers to complete.", "", "", "", "", "", "", "", ""],
    ["khufu_3", "great_pyramid", 3, "fact", "", "", "The pyramid contains about 2.3 million stone blocks, each weighing 2.5 tons on average!", "", "", "", "", "", "", ""],
    ["khufu_4", "great_pyramid", 4, "quiz", "", "", "", "Who was the Great Pyramid built for?", "Pharaoh Khufu", "Pharaoh Tutankhamun", "Cleopatra", "Ramesses II", 1, "The Great Pyramid was built as a tomb for Pharaoh Khufu (also known as Cheops) of the Fourth Dynasty."],
    ["khufu_5", "great_pyramid", 5, "story", "", "Inside are three chambers: the King's Chamber with a granite sarcophagus, the Queen's Chamber, and an unfinished underground chamber.", "", "", "", "", "", "", "", ""],
    ["khufu_6", "great_pyramid", 6, "fact", "", "", "The four sides of the pyramid are aligned almost perfectly with the four cardinal directions!", "", "", "", "", "", "", ""],

    # Sphinx cards
    ["sphinx_1", "sphinx", 1, "intro", "", "The Great Sphinx is the oldest and largest monolith statue in the world, carved from a single limestone bedrock.", "", "", "", "", "", "", "", ""],
    ["sphinx_2", "sphinx", 2, "story", "", "Standing 66 feet high and 240 feet long, the Sphinx has the body of a lion and the face believed to be of Pharaoh Khafre.", "", "", "", "", "", "", "", ""],
    ["sphinx_3", "sphinx", 3, "fact", "", "", "The Sphinx's nose is missing! Contrary to legend, it wasn't shot off by Napoleon's soldiers.", "", "", "", "", "", "", ""],
    ["sphinx_4", "sphinx", 4, "quiz", "", "", "", "What does 'Abu al-Hol' (Arabic name) mean?", "Father of Terror", "Guardian of the Dead", "Sun God", "Stone Lion", 1, "The Arabic name 'Abu al-Hol' translates to 'Father of Terror' or 'Father of Dread'."],

    # Khafre Pyramid cards
    ["khafre_1", "khafre_pyramid", 1, "intro", "", "Though slightly smaller than the Great Pyramid, Khafre's pyramid appears taller because it's built on higher ground.", "", "", "", "", "", "", "", ""],
    ["khafre_2", "khafre_pyramid", 2, "story", "", "It's the only pyramid that still has some of its original white limestone casing at the top, giving us a glimpse of how all pyramids once gleamed.", "", "", "", "", "", "", "", ""],
    ["khafre_3", "khafre_pyramid", 3, "fact", "", "", "The pyramids were originally covered in polished white limestone that would have shone brilliantly in the sun!", "", "", "", "", "", "", ""],

    # Solar Boat cards
    ["boat_1", "solar_boat", 1, "intro", "", "In 1954, archaeologists discovered a 4,600-year-old boat buried in a pit beside the Great Pyramid.", "", "", "", "", "", "", "", ""],
    ["boat_2", "solar_boat", 2, "story", "", "The boat was found in 1,224 pieces. It took 14 years to carefully reassemble this ancient vessel.", "", "", "", "", "", "", "", ""],
    ["boat_3", "solar_boat", 3, "fact", "", "", "At 143 feet long, it's one of the oldest and largest ancient vessels ever found!", "", "", "", "", "", "", ""],
    ["boat_4", "solar_boat", 4, "quiz", "", "", "", "What was the solar boat's purpose?", "Carry the pharaoh's soul to the afterlife", "Fishing on the Nile", "Military transport", "Trade with other kingdoms", 1, "The solar boat was meant to carry the pharaoh's soul across the heavens with the sun god Ra."],

    # Luxor Temple cards
    ["luxor_e1", "luxor_entrance", 1, "intro", "", "Two massive seated statues of Ramesses II guard the entrance, each standing 50 feet tall.", "", "", "", "", "", "", "", ""],
    ["luxor_e2", "luxor_entrance", 2, "story", "", "Originally, two obelisks stood here. One remains in Luxor; the other has been in Paris since 1836.", "", "", "", "", "", "", "", ""],
    ["luxor_e3", "luxor_entrance", 3, "fact", "", "", "The missing obelisk now stands in Place de la Concorde in Paris - a gift from Egypt to France!", "", "", "", "", "", "", ""],

    # Tutankhamun cards
    ["tut_1", "tut_tomb", 1, "intro", "", "In 1922, Howard Carter made the discovery of a lifetime - the nearly intact tomb of a young pharaoh.", "", "", "", "", "", "", "", ""],
    ["tut_2", "tut_tomb", 2, "story", "", "\"Can you see anything?\" Lord Carnarvon asked. \"Yes, wonderful things,\" Carter whispered, peering through a small hole.", "", "", "", "", "", "", "", ""],
    ["tut_3", "tut_tomb", 3, "fact", "", "", "It took Carter 10 years to catalog all 5,398 objects found in the tomb!", "", "", "", "", "", "", ""],
    ["tut_4", "tut_tomb", 4, "quiz", "", "", "", "Why was Tutankhamun's tomb so significant?", "It was nearly intact", "It was the largest", "It was the oldest", "It had the most mummies", 1, "Unlike most royal tombs which were robbed in antiquity, Tut's tomb was found nearly intact with all its treasures."],

    # Karnak cards
    ["hypo_1", "hypostyle_hall", 1, "intro", "", "You stand among giants. 134 massive columns surround you, the largest standing 69 feet tall.", "", "", "", "", "", "", "", ""],
    ["hypo_2", "hypostyle_hall", 2, "story", "", "Imagine this space 3,000 years ago: columns painted in brilliant colors, incense smoke drifting through light.", "", "", "", "", "", "", "", ""],
    ["hypo_3", "hypostyle_hall", 3, "fact", "", "", "The entire Notre-Dame cathedral could fit inside the Hypostyle Hall!", "", "", "", "", "", "", ""],

    # Abu Simbel cards
    ["abu_1", "great_temple", 1, "intro", "", "Four massive statues of Ramesses II gaze eternally across the Nubian desert, each 65 feet tall.", "", "", "", "", "", "", "", ""],
    ["abu_2", "great_temple", 2, "story", "", "Twice a year, on Feb 22 and Oct 22, the rising sun penetrates 185 feet into the mountain to illuminate the inner statues.", "", "", "", "", "", "", "", ""],
    ["abu_3", "great_temple", 3, "fact", "", "", "The entire temple was moved in the 1960s to save it from flooding! Cut into blocks and reassembled on higher ground.", "", "", "", "", "", "", ""],
    ["abu_4", "great_temple", 4, "quiz", "", "", "", "Why was Abu Simbel moved?", "To save it from Lake Nasser flooding", "It was sinking", "For tourism access", "Political reasons", 1, "The Aswan High Dam would have submerged the temple. UNESCO led an international effort to relocate it."],
]

for row in cards_data:
    ws_cards.append(row)

style_header(ws_cards, len(cards_headers))
auto_width(ws_cards)

# Add dropdown for card type
type_values = '"intro,story,fact,quiz,summary"'
dv_type = DataValidation(type="list", formula1=type_values, allow_blank=True)
ws_cards.add_data_validation(dv_type)
dv_type.add('D2:D200')

# ============ SHEET 4: Tips ============
ws_tips = wb.create_sheet("Tips")

tips_headers = ["id", "siteId", "order", "tip"]
ws_tips.append(tips_headers)

tips_data = [
    ["giza_tip_1", "giza", 1, "Bring plenty of water and sun protection"],
    ["giza_tip_2", "giza", 2, "Hire an official guide at the entrance"],
    ["giza_tip_3", "giza", 3, "The interior of the Great Pyramid is hot and cramped"],
    ["giza_tip_4", "giza", 4, "Visit the Solar Boat Museum"],
    ["luxor_tip_1", "luxor", 1, "Visit at sunset to see the temple transform"],
    ["luxor_tip_2", "luxor", 2, "The temple is in central Luxor - easy walk from hotels"],
    ["luxor_tip_3", "luxor", 3, "Look for the Abu Haggag Mosque built on top"],
    ["valley_tip_1", "valley_kings", 1, "Your ticket includes 3 tombs - choose wisely"],
    ["valley_tip_2", "valley_kings", 2, "Tutankhamun's tomb requires a separate ticket"],
    ["valley_tip_3", "valley_kings", 3, "No photography inside the tombs"],
    ["karnak_tip_1", "karnak", 1, "The complex is vast - wear comfortable shoes"],
    ["karnak_tip_2", "karnak", 2, "Don't miss the Sacred Lake"],
    ["karnak_tip_3", "karnak", 3, "The Hypostyle Hall is best in morning light"],
    ["abu_simbel_tip_1", "abu_simbel", 1, "Most visitors fly from Aswan"],
    ["abu_simbel_tip_2", "abu_simbel", 2, "The smaller temple honors Nefertari"],
    ["abu_simbel_tip_3", "abu_simbel", 3, "Visit the UNESCO relocation exhibition"],
]

for row in tips_data:
    ws_tips.append(row)

style_header(ws_tips, len(tips_headers))
auto_width(ws_tips)

# ============ SHEET 5: ArabicPhrases ============
ws_phrases = wb.create_sheet("ArabicPhrases")

phrases_headers = ["id", "siteId", "english", "arabic", "pronunciation"]
ws_phrases.append(phrases_headers)

phrases_data = [
    ["giza_phrase_1", "giza", "Pyramid", "هرم", "haram"],
    ["giza_phrase_2", "giza", "How much?", "بكام؟", "bekam?"],
    ["giza_phrase_3", "giza", "Thank you", "شكراً", "shukran"],
    ["luxor_phrase_1", "luxor", "Temple", "معبد", "ma'bad"],
    ["luxor_phrase_2", "luxor", "Beautiful", "جميل", "gameel"],
    ["valley_phrase_1", "valley_kings", "Tomb", "مقبرة", "maqbara"],
    ["valley_phrase_2", "valley_kings", "King", "ملك", "malik"],
    ["karnak_phrase_1", "karnak", "Column", "عمود", "amood"],
    ["karnak_phrase_2", "karnak", "Huge", "كبير", "kebeer"],
    ["abu_simbel_phrase_1", "abu_simbel", "Sun", "شمس", "shams"],
    ["abu_simbel_phrase_2", "abu_simbel", "Amazing", "مذهل", "muzhil"],
]

for row in phrases_data:
    ws_phrases.append(row)

style_header(ws_phrases, len(phrases_headers))
auto_width(ws_phrases)

# Save the workbook
output_path = "/Users/nareman/Documents/Projects/UnlockEgypt/ContentManagement/UnlockEgypt_Content.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
print(f"\nSheets created:")
print(f"  - Sites: 5 rows")
print(f"  - SubLocations: 8 rows")
print(f"  - Cards: 32 rows")
print(f"  - Tips: 16 rows")
print(f"  - ArabicPhrases: 11 rows")
print(f"\nDropdowns added for: era, tourismType, placeType, city, card type")
